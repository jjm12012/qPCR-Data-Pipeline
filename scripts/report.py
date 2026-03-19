"""
report.py  –  Step 4 of the qPCR pipeline
Compiles all processed data, statistics, and figure references
into a formatted Excel workbook: qpcr_summary_report.xlsx

Sheets:
  Summary           – Key statistics table + run metadata
  Per-Run Detail    – Full parsed data for every run (conditional formatting for QC)
  QC Flags          – All flagged wells with red row highlighting
  Sample Type Stats – Detailed statistics per sample type
  Run Comparison    – Inter-run positivity pivot table
  Controls          – Plate control statistics
  Notes             – Pipeline parameters and interpretation guide
"""

import argparse
import base64
from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _hdr_fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_colour, end_color=hex_colour)

def _font(bold=False, size=10, colour="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=colour)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL      = _hdr_fill("1F497D")
SUBHDR_FILL      = _hdr_fill("4472C4")
ALT_FILL         = _hdr_fill("EEF3FA")
POSITIVE_FILL    = _hdr_fill("C6EFCE")
NEGATIVE_FILL    = _hdr_fill("FFC7CE")
WHITE_FILL       = _hdr_fill("FFFFFF")
BORDERLINE_FILL  = _hdr_fill("FFFF00")   # yellow – Borderline Cq
IC_FAIL_FILL     = _hdr_fill("FFC7CE")   # light red – IC extraction failure
INHIBITION_FILL  = _hdr_fill("FCE4B0")   # orange – Inhibition result
QC_ROW_FILL      = _hdr_fill("FFC7CE")   # light red – QC Flags sheet rows
QC_HDR_FILL      = _hdr_fill("C00000")   # dark red – QC Flags sheet header


def _write_df_to_sheet(ws, df: pd.DataFrame,
                       start_row: int = 1, start_col: int = 1,
                       header_colour: str = "1F497D"):
    """Write a DataFrame to a worksheet with formatted headers."""
    hdr_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    hdr_fill  = _hdr_fill(header_colour)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    for ci, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _thin_border()

    # Data rows
    for ri, data_row in enumerate(dataframe_to_rows(df, index=False, header=False),
                                  start=start_row + 1):
        for ci, value in enumerate(data_row, start=start_col):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = _font()
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Alternating row fill
            if (ri - start_row) % 2 == 0:
                cell.fill = ALT_FILL


def _auto_col_widths(ws, min_width=10, max_width=40):
    for col in ws.columns:
        width = min_width
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                width = min(max(width, cell_len + 2), max_width)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_summary(wb: Workbook, stats: pd.DataFrame,
                   combined: pd.DataFrame, controls: pd.DataFrame | None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 30
    title_cell = ws["A1"]
    title_cell.value     = "qPCR Automated Pipeline — Summary Report"
    title_cell.font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = _center()
    ws.merge_cells("A1:I1")

    # Generated timestamp
    ts_cell = ws["A2"]
    ts_cell.value     = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ts_cell.font      = _font(size=9, colour="666666")
    ts_cell.alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:I2")

    # Key metrics row
    n_runs     = combined["run_id"].nunique()
    n_samples  = len(combined[~combined["is_control"]])
    n_positive = (combined[~combined["is_control"]]["result"]
                  .str.strip().str.lower() == "positive").sum()
    overall_rate = n_positive / n_samples * 100 if n_samples else 0

    metrics = [
        ("Runs Processed", str(n_runs)),
        ("Sample Wells",   str(n_samples)),
        ("Overall Positive Rate", f"{overall_rate:.1f}%"),
        ("Sample Types", str(stats["sample_type"].nunique())),
    ]
    ws.row_dimensions[4].height = 22
    for ci, (label, val) in enumerate(metrics, start=1):
        label_cell = ws.cell(row=4, column=ci * 2 - 1, value=label)
        val_cell   = ws.cell(row=4, column=ci * 2,     value=val)
        label_cell.font  = _font(bold=True, size=10, colour="FFFFFF")
        label_cell.fill  = SUBHDR_FILL
        label_cell.alignment = _center()
        val_cell.font    = _font(bold=True, size=12)
        val_cell.alignment = _center()

    # Sample-type summary table
    ws.cell(row=6, column=1, value="Sample Type Statistics").font = _font(bold=True, size=12)

    display_stats = stats[[
        "sample_type", "n_wells", "n_positive", "n_negative",
        "positivity_rate_pct", "mean_target_cq", "sd_target_cq",
        "mean_ic_cq", "sd_ic_cq",
    ]].copy()
    display_stats.columns = [
        "Sample Type", "N Wells", "N Positive", "N Negative",
        "Positivity %", "Mean Target Cq", "SD Target Cq",
        "Mean I.C. Cq", "SD I.C. Cq",
    ]
    _write_df_to_sheet(ws, display_stats, start_row=7, header_colour="4472C4")
    _auto_col_widths(ws)


def _sheet_detail(wb: Workbook, combined: pd.DataFrame, qc_flags=None):
    ws = wb.create_sheet("Per-Run Detail")
    ws.sheet_view.showGridLines = False

    display = combined[[
        "run_id", "run_date", "assay_name", "lot_number", "extraction_lot",
        "well", "content", "sample_label", "sample_type", "replicate",
        "target_cq", "ic_cq", "result",
    ]].copy()
    display.columns = [
        "Run ID", "Run Date", "Assay Name", "Lot Number", "Extraction Lot",
        "Well", "Content", "Sample Label", "Sample Type", "Replicate",
        "Target Cq", "I.C. Cq", "Result",
    ]
    _write_df_to_sheet(ws, display, start_row=1)

    # Build lookup sets from qc_flags for conditional formatting
    ic_fail_wells = set()
    if qc_flags is not None and not qc_flags.empty:
        ic_rows = qc_flags[qc_flags["flag_reason"].str.contains("IC extraction", na=False)]
        for _, r in ic_rows.iterrows():
            ic_fail_wells.add((str(r["run_id"]), str(r["well"])))

    col_names = list(display.columns)
    result_col   = col_names.index("Result") + 1
    target_cq_col = col_names.index("Target Cq") + 1
    ic_cq_col    = col_names.index("I.C. Cq") + 1
    run_id_col   = col_names.index("Run ID") + 1
    well_col     = col_names.index("Well") + 1

    for row_idx in range(2, len(display) + 2):
        result_cell = ws.cell(row=row_idx, column=result_col)
        cq_cell     = ws.cell(row=row_idx, column=target_cq_col)
        ic_cell     = ws.cell(row=row_idx, column=ic_cq_col)
        run_cell    = ws.cell(row=row_idx, column=run_id_col)
        well_cell   = ws.cell(row=row_idx, column=well_col)

        result_val = str(result_cell.value).strip().lower()

        # Result colouring
        if result_val == "positive":
            result_cell.fill = POSITIVE_FILL
        elif result_val == "negative":
            result_cell.fill = NEGATIVE_FILL
        elif result_val == "inhibition":
            result_cell.fill = INHIBITION_FILL

        # Yellow fill: Borderline Cq (Target Cq > 35)
        try:
            cq_val = float(cq_cell.value)
            if cq_val > 35:
                cq_cell.fill = BORDERLINE_FILL
        except (TypeError, ValueError):
            pass

        # Light-red fill: IC extraction failure
        key = (str(run_cell.value), str(well_cell.value))
        if key in ic_fail_wells:
            ic_cell.fill = IC_FAIL_FILL

    # Freeze header row
    ws.freeze_panes = "A2"
    _auto_col_widths(ws)


def _sheet_qc_flags(wb: Workbook, qc_flags):
    """QC Flags sheet: all flagged wells with red row highlighting."""
    if qc_flags is None or (hasattr(qc_flags, "empty") and qc_flags.empty):
        return

    ws = wb.create_sheet("QC Flags")
    ws.sheet_view.showGridLines = False

    display = qc_flags[["run_id", "well", "sample_type", "flag_reason"]].copy()
    display.columns = ["Run ID", "Well", "Sample Type", "Flag Reason"]

    hdr_font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row with dark red fill
    for ci, col in enumerate(display.columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = hdr_font
        cell.fill      = QC_HDR_FILL
        cell.alignment = hdr_align
        cell.border    = _thin_border()

    # Data rows with light red fill on every row
    for ri, data_row in enumerate(dataframe_to_rows(display, index=False, header=False),
                                  start=2):
        for ci, value in enumerate(data_row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = _font()
            cell.fill      = QC_ROW_FILL
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    _auto_col_widths(ws)


def _sheet_stats(wb: Workbook, stats: pd.DataFrame):
    ws = wb.create_sheet("Sample Type Stats")
    ws.sheet_view.showGridLines = False
    _write_df_to_sheet(ws, stats, start_row=1)
    _auto_col_widths(ws)


def _sheet_plate_layout(wb: Workbook, combined: pd.DataFrame):
    """
    Plate Layout sheet: one sub-grid per run showing each well's sample ID,
    coloured by result (same scheme as the plate heatmap figure).
    """
    result_fills = {
        "positive":   _hdr_fill("FFC7CE"),   # red
        "negative":   _hdr_fill("C6EFCE"),   # green
        "valid ctrl": _hdr_fill("BDD7EE"),   # blue
        "inhibition": _hdr_fill("FCE4B0"),   # orange
    }
    empty_fill    = _hdr_fill("F2F2F2")
    row_label_fill = _hdr_fill("D9E1F2")

    ws = wb.create_sheet("Plate Layout")
    ws.sheet_view.showGridLines = False

    ROWS = list("ABCDEFGH")
    COLS = list(range(1, 13))

    cursor_row = 1   # current write position in the sheet

    for run_id in sorted(combined["run_id"].unique()):
        sub = combined[combined["run_id"] == run_id].copy()

        # Build lookup: (row_letter, col_num) → (sample_label, result)
        well_map = {}
        for _, r in sub.iterrows():
            m = re.match(r"([A-Ha-h])(\d+)", str(r["well"]).strip())
            if not m:
                continue
            letter = m.group(1).upper()
            col_n  = int(m.group(2))
            label  = str(r["sample_label"]).strip()
            if label in ("", "nan"):
                label = str(r["content"]).strip()
            result = str(r["result"]).strip().lower()
            well_map[(letter, col_n)] = (label, result)

        # Run title
        title_cell = ws.cell(row=cursor_row, column=1, value=f"Run: {run_id}")
        title_cell.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        title_cell.fill      = HEADER_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=cursor_row, start_column=1,
            end_row=cursor_row,   end_column=13
        )
        ws.row_dimensions[cursor_row].height = 18
        cursor_row += 1

        # Column headers (blank corner + 01–12)
        corner = ws.cell(row=cursor_row, column=1, value="")
        corner.fill = row_label_fill
        for ci, col_n in enumerate(COLS, start=2):
            cell = ws.cell(row=cursor_row, column=ci, value=f"{col_n:02d}")
            cell.font      = Font(name="Calibri", bold=True, size=9)
            cell.fill      = row_label_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()
        ws.row_dimensions[cursor_row].height = 16
        cursor_row += 1

        # Data rows A–H
        for row_letter in ROWS:
            # Row label
            lbl_cell = ws.cell(row=cursor_row, column=1, value=row_letter)
            lbl_cell.font      = Font(name="Calibri", bold=True, size=9)
            lbl_cell.fill      = row_label_fill
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.border    = _thin_border()

            for ci, col_n in enumerate(COLS, start=2):
                cell = ws.cell(row=cursor_row, column=ci)
                key  = (row_letter, col_n)
                if key in well_map:
                    sample_label, result = well_map[key]
                    cell.value     = sample_label
                    cell.fill      = result_fills.get(result, empty_fill)
                else:
                    cell.value = ""
                    cell.fill  = empty_fill
                cell.font      = Font(name="Calibri", size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = _thin_border()
            ws.row_dimensions[cursor_row].height = 36
            cursor_row += 1

        cursor_row += 2   # blank rows between runs

    # Fixed column widths
    ws.column_dimensions["A"].width = 5   # row label
    for ci in range(2, 14):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _sheet_run_comparison(wb: Workbook, run_comparison):
    """Run Comparison sheet: inter-run positivity pivot."""
    if run_comparison is None or (hasattr(run_comparison, "empty") and run_comparison.empty):
        return

    ws = wb.create_sheet("Run Comparison")
    ws.sheet_view.showGridLines = False
    _write_df_to_sheet(ws, run_comparison, start_row=1, header_colour="7030A0")
    _auto_col_widths(ws)


def _sheet_controls(wb: Workbook, controls: pd.DataFrame):
    if controls is None or controls.empty:
        return
    ws = wb.create_sheet("Controls")
    ws.sheet_view.showGridLines = False
    _write_df_to_sheet(ws, controls, start_row=1, header_colour="7030A0")
    _auto_col_widths(ws)


def _sheet_notes(wb: Workbook):
    ws = wb.create_sheet("Notes")
    ws.sheet_view.showGridLines = False

    notes = [
        ("qPCR Pipeline — Interpretation Guide", True, 14),
        ("", False, 10),
        ("COLUMN DEFINITIONS", True, 11),
        ("Target Cq:  Quantification cycle for the target gene. Lower = more template present.", False, 10),
        ("I.C. Cq:    Internal Control Cq. Indicates extraction efficiency. Should be consistent across sample types.", False, 10),
        ("Result:     Positive = target detected above threshold.  Negative = target not detected.  Valid Ctrl = control passed.", False, 10),
        ("Positivity %: Percentage of wells within each sample type that returned a Positive result.", False, 10),
        ("", False, 10),
        ("QC FLAG DEFINITIONS", True, 11),
        ("Borderline (yellow):        Target Cq > 35 — result may be near the detection limit.", False, 10),
        ("IC extraction failure (red): I.C. Cq >2 SD from the plate mean — possible extraction problem.", False, 10),
        ("", False, 10),
        ("SAMPLE TYPE DEFINITIONS (auto-parsed from sample labels)", True, 11),
        ("Manual_Pos:           Manually prepared positive samples.", False, 10),
        ("Manual_Neg:           Manually prepared negative samples.", False, 10),
        ("Auto_Pos:             Automatically extracted positive samples.", False, 10),
        ("Auto_Neg:             Automatically extracted negative samples.", False, 10),
        ("Lysis_5mL_Control:    Lysis buffer process control at 5 mL.", False, 10),
        ("Lysis_1.25mL_Control: Lysis buffer process control at 1.25 mL.", False, 10),
        ("", False, 10),
        ("PIPELINE STEPS", True, 11),
        ("1. parse_raw.py    — Converts raw .xlsx to normalised CSV; extracts metadata.", False, 10),
        ("2. analyze.py      — Computes statistics, QC flags, and inter-run CV% across all runs.", False, 10),
        ("3. visualize.py    — Generates 9 figure types (see results/figures/).", False, 10),
        ("4. report.py       — Compiles this workbook.", False, 10),
    ]

    for ri, (text, bold, size) in enumerate(notes, start=1):
        cell = ws.cell(row=ri, column=1, value=text)
        cell.font = Font(name="Calibri", bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 90


# ---------------------------------------------------------------------------
# HTML figure report
# ---------------------------------------------------------------------------

# Human-readable captions keyed by filename prefix
_FIG_CAPTIONS = {
    "01_positivity_rate_bar":    "Positivity Rate by Sample Type",
    "04_cq_scatter_by_well":     "Target Cq vs. I.C. Cq Scatter",
    "05_plate_heatmap_result":   "Plate Result Map",
    "06_plate_heatmap_cq":       "Plate Target Cq Heatmap",
    "07_run_positivity_heatmap": "Run × Sample-Type Positivity Heatmap",
    "08_ic_cq_control_chart":    "Cq Spatial & Sample-Type Overview",
    "09_interrun_cv_bar":        "Inter-run CV% by Sample Type",
}


def _fig_caption(stem: str) -> str:
    """Return a caption for a figure file stem, falling back to the stem itself."""
    for prefix, caption in _FIG_CAPTIONS.items():
        if stem.startswith(prefix):
            return caption
    return stem


def _write_html_report(fig_dir: Path, out_dir: Path,
                       stats: pd.DataFrame, combined: pd.DataFrame,
                       qc_flags, timestamp: str) -> Path:
    """
    Write a self-contained HTML file with embedded figures and a summary table.
    Images are base64-encoded so the file can be opened anywhere without
    needing the figures/ folder alongside it.
    """
    figs = sorted(fig_dir.glob("*.png"))

    # --- Key metrics ---
    n_runs      = combined["run_id"].nunique()
    n_samples   = len(combined[~combined["is_control"]])
    n_positive  = int((combined[~combined["is_control"]]["result"]
                       .str.strip().str.lower() == "positive").sum())
    rate        = round(n_positive / n_samples * 100, 1) if n_samples else 0
    n_flags     = len(qc_flags) if qc_flags is not None and not qc_flags.empty else 0

    # --- Stats table rows ---
    stats_rows = ""
    for _, row in stats.iterrows():
        n_pos  = int(row["n_positive"])
        n_tot  = int(row["n_wells"])
        pct    = float(row["positivity_rate_pct"])
        mean_cq = f"{row['mean_target_cq']:.3f}" if row["mean_target_cq"] == row["mean_target_cq"] else "—"
        sd_cq   = f"{row['sd_target_cq']:.3f}"   if row["sd_target_cq"]   == row["sd_target_cq"]   else "—"
        stats_rows += (
            f"<tr><td>{row['sample_type']}</td>"
            f"<td>{n_pos}/{n_tot}</td>"
            f"<td>{pct:.1f}%</td>"
            f"<td>{mean_cq}</td>"
            f"<td>{sd_cq}</td></tr>\n"
        )

    # --- QC flags table ---
    if n_flags:
        flag_rows = ""
        for _, r in qc_flags.iterrows():
            flag_rows += (
                f"<tr><td>{r['run_id']}</td><td>{r['well']}</td>"
                f"<td>{r['sample_type']}</td><td>{r['flag_reason']}</td></tr>\n"
            )
        qc_section = f"""
        <section>
          <h2 style="color:#c00000;">&#9888; QC Flags ({n_flags})</h2>
          <table>
            <thead><tr>
              <th>Run ID</th><th>Well</th><th>Sample Type</th><th>Flag Reason</th>
            </tr></thead>
            <tbody>{flag_rows}</tbody>
          </table>
        </section>"""
    else:
        qc_section = """
        <section>
          <h2 style="color:#2e7d32;">&#10003; No QC Flags</h2>
        </section>"""

    # --- Figure cards ---
    fig_cards = ""
    for fig_path in figs:
        b64 = base64.b64encode(fig_path.read_bytes()).decode("ascii")
        caption = _fig_caption(fig_path.stem)
        fig_cards += f"""
        <figure>
          <img src="data:image/png;base64,{b64}" alt="{caption}">
          <figcaption>{caption}</figcaption>
        </figure>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>qPCR Pipeline Report</title>
  <style>
    body {{
      font-family: Calibri, Arial, sans-serif;
      margin: 0; padding: 20px 40px;
      background: #f5f7fa; color: #222;
    }}
    h1 {{ background: #1f497d; color: #fff; padding: 16px 24px;
          margin: 0 -40px 24px; font-size: 1.6em; }}
    h2 {{ color: #1f497d; border-bottom: 2px solid #4472c4;
          padding-bottom: 4px; margin-top: 32px; }}
    .meta {{ color: #666; font-size: 0.85em; margin-bottom: 24px; }}
    .metrics {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
    .metric-card {{
      background: #fff; border: 1px solid #d0d8e8; border-radius: 6px;
      padding: 14px 20px; min-width: 140px; text-align: center;
    }}
    .metric-card .value {{ font-size: 1.8em; font-weight: bold; color: #1f497d; }}
    .metric-card .label {{ font-size: 0.8em; color: #666; margin-top: 4px; }}
    section {{ background: #fff; border: 1px solid #d0d8e8; border-radius: 6px;
               padding: 20px 24px; margin-bottom: 24px; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 0.9em; }}
    th {{ background: #4472c4; color: #fff; padding: 8px 12px; text-align: left; }}
    td {{ padding: 7px 12px; border-bottom: 1px solid #e0e6f0; }}
    tr:nth-child(even) td {{ background: #eef3fa; }}
    figure {{
      background: #fff; border: 1px solid #d0d8e8; border-radius: 6px;
      padding: 16px; margin: 0 0 20px; text-align: center;
    }}
    figure img {{ max-width: 100%; height: auto; }}
    figcaption {{ margin-top: 8px; font-size: 0.85em; color: #555; }}
  </style>
</head>
<body>
  <h1>qPCR Automated Pipeline &mdash; Figure Report</h1>
  <p class="meta">Generated: {timestamp} &nbsp;|&nbsp; {n_runs} run(s) processed</p>

  <div class="metrics">
    <div class="metric-card">
      <div class="value">{n_samples}</div>
      <div class="label">Sample Wells</div>
    </div>
    <div class="metric-card">
      <div class="value">{n_positive}/{n_samples}</div>
      <div class="label">Positive</div>
    </div>
    <div class="metric-card">
      <div class="value">{rate}%</div>
      <div class="label">Positivity Rate</div>
    </div>
    <div class="metric-card">
      <div class="value" style="color:{'#c00000' if n_flags else '#2e7d32'}">{n_flags}</div>
      <div class="label">QC Flags</div>
    </div>
  </div>

  <section>
    <h2>Sample Type Summary</h2>
    <table>
      <thead><tr>
        <th>Sample Type</th><th>Positive</th><th>Positivity %</th>
        <th>Mean Target Cq</th><th>SD Target Cq</th>
      </tr></thead>
      <tbody>{stats_rows}</tbody>
    </table>
  </section>

  {qc_section}

  <h2>Figures</h2>
  {fig_cards}

</body>
</html>"""

    out_path = out_dir / "qpcr_figures_report.html"
    out_path.write_text(html, encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Compile qPCR Excel summary report.")
    parser.add_argument("--proc_dir",   required=True)
    parser.add_argument("--stats",      required=True)
    parser.add_argument("--fig_dir",    required=True)
    parser.add_argument("--output_dir", required=True)
    args = parser.parse_args()

    proc_dir = Path(args.proc_dir)
    out_dir  = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load data
    files    = sorted(proc_dir.glob("*_parsed.csv"))
    combined = pd.concat([pd.read_csv(f) for f in files], ignore_index=True)
    stats    = pd.read_csv(args.stats)

    ctrl_path = out_dir / "control_statistics.csv"
    controls  = pd.read_csv(ctrl_path) if ctrl_path.exists() else None

    qc_flags_path = out_dir / "qc_flags.csv"
    qc_flags = pd.read_csv(qc_flags_path) if qc_flags_path.exists() else None

    run_cmp_path = out_dir / "run_comparison.csv"
    run_comparison = pd.read_csv(run_cmp_path) if run_cmp_path.exists() else None

    # Build workbook — sheet order:
    # Summary → Per-Run Detail → QC Flags → Sample Type Stats →
    # Run Comparison → Controls → Notes
    wb = Workbook()
    wb.remove(wb.active)   # Remove default empty sheet

    _sheet_summary(wb, stats, combined, controls)
    _sheet_detail(wb, combined, qc_flags=qc_flags)
    _sheet_plate_layout(wb, combined)
    _sheet_qc_flags(wb, qc_flags)
    _sheet_stats(wb, stats)
    _sheet_run_comparison(wb, run_comparison)
    _sheet_controls(wb, controls)
    _sheet_notes(wb)

    out_path = out_dir / "qpcr_summary_report.xlsx"
    wb.save(out_path)
    print(f"  [report] Report saved → {out_path}")

    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    html_path = _write_html_report(
        fig_dir   = Path(args.fig_dir),
        out_dir   = out_dir,
        stats     = stats,
        combined  = combined,
        qc_flags  = qc_flags,
        timestamp = timestamp,
    )
    print(f"  [report] HTML figures report → {html_path}")


if __name__ == "__main__":
    main()
